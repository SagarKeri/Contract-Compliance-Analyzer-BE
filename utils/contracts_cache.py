from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

def format_sheet(writer, sheet_name: str):
    worksheet = writer.sheets[sheet_name]
    
    # Define color styles
    metadata_fill = PatternFill(start_color="0F6368", end_color="0F6368", fill_type="solid")
    table_fill = PatternFill(start_color="299E66", end_color="299E66", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)  # White and bold for headers
    white_font_regular = Font(color="FFFFFF")  # White for regular text
    
    # Apply formatting to metadata section (rows 1-3)
    for row in worksheet[1:3]:  # Rows are 1-based in openpyxl
        for cell in row:
            cell.fill = metadata_fill
            cell.font = white_font_regular
    
    # Apply formatting to table header (row 5, since startrow=4 in df_table.to_excel)
    for cell in worksheet[5]:  # Row 5 is the table header
        cell.fill = table_fill
        cell.font = white_font  # Bold white font for headers
    
    # Apply formatting to table data (rows 6 and below)
    for row in worksheet.iter_rows(min_row=6, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.fill = table_fill
            cell.font = white_font_regular
    
    # Adjust column widths based on content
    for col_idx, col in enumerate(worksheet.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = max_length + 2
