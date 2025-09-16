import os
import json
from pathlib import Path
import fitz  # PyMuPDF
from fastapi import FastAPI, File, UploadFile, HTTPException, Form, Query
from fastapi.responses import JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from langchain_community.llms import Ollama
import google.generativeai as genai
import hashlib
from pathlib import Path
import json
from pydantic import BaseModel
from typing import Optional
from pymongo import MongoClient
from io import BytesIO
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_community.vectorstores import FAISS
from langchain_huggingface import HuggingFaceEmbeddings


# ---------- Database Connection ----------
client = MongoClient("mongodb://localhost:27017/")
db = client["Contract-Compliance-Analyzer"]

# ---------- Load Gemini API Key ---------- #
genai.configure(api_key=os.getenv("GEMINI_API_KEY", "AIzaSyDQs8IFeuZvCd_erNBrEXU2Q8rqaCUG-pc"))

# ---------- Initialize FastAPI App ---------- #
app = FastAPI(
    title="Oil & Gas Contract Compliance API",
    description="Upload oil & gas contracts in PDF format and get compliance analysis across 19 regulatory areas.",
    version="1.6.0",
)

from fastapi.middleware.cors import CORSMiddleware

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # or your frontend URL
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],  
)


#---------------load compliances from the mongo DB-----------#
from typing import List
from fastapi import HTTPException
from pymongo import MongoClient

# Existing MongoDB connection
# client = MongoClient("mongodb://localhost:27017/")
# db = client["Contract-Compliance-Analyzer"]
clauses_collection = db["clauses"] 
user_analysis = db["user_analysis"]


from pymongo.errors import PyMongoError

def load_selected_clauses(clause_ids: list[int]) -> list:
    if not clause_ids:
        raise HTTPException(status_code=400, detail="No clause IDs provided.")

    try:
        # Convert to int if stored as int, or to str if stored as string
        clauses = list(
            clauses_collection.find(
                {"_id": {"$in": clause_ids}},  # match your real field type
                {"_id": 0, "clause_name": 1, "clause_text": 1}
            )
        )

        if not isinstance(clauses, list):
            clauses = [clauses]  # ensure always a list

        if not clauses:
            raise HTTPException(status_code=404, detail="No clauses found for given IDs.")

        return clauses

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

# ---------- PDF Text Extractor ---------- #
def extract_text_from_pdf(file_bytes: bytes) -> str:
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        text = ""
        for page in doc:
            # Use "text" option to preserve line breaks and whitespace
            page_text = page.get_text("text")
            text += page_text + "\n"  # Add page separator
        return text.strip()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"PDF processing failed: {str(e)}")

# ---------- Prompt Template ---------- #
def build_prompt(summary_context: str, clauses_section: str, model: str) -> str:
    """
    Build a prompt to:
    1. Generate a brief description of the uploaded contract PDF based on summary context.
    2. Check if each clause exists in the contract using provided relevant excerpts.
    """
    return f"""
    You are an expert contract reviewer.

    Task 1: Provide a brief description of the uploaded contract based on the following excerpt: 
    {summary_context}
    - The description MUST be at least 50 words long.
    - It should be written in 2–3 complete sentences.
    - Cover key parties, purpose, subject matter, and type of obligations.

    Task 2: Analyze the contract and check for the presence and adequacy of the following clauses using their respective relevant excerpts:

    {clauses_section}

    ### Instructions:
    - For each clause:
        - Use only the provided relevant excerpts for that clause.
        - If it is completely missing, return:
            "missing_clause": "Missing",
            "reason":"<brief reason>",
            "extracted_text": null
        - If it exists but is vague, generic, or non-compliant, return:
            "missing_clause": "Insufficient",
            "reason":"<brief reason>",
            "extracted_text": "<first 50 words from the original clause in the contract>"
        - If it exists and is fully adequate, return:
            "missing_clause": "Sufficient",
            "reason":"<brief reason>",
            "extracted_text": "<first 50 words from the original clause in the contract>"

    - Preserve exact formatting from the contract in "extracted_text".
    - Output a single JSON object in the format:

    {{
      "description": "<2–3 sentence summary of the contract>",
      "analysis": [
        {{
          "compliance_area": "<clause_name>",
          "missing_clause": "<Missing / Insufficient / Sufficient>",
          "reason": "<brief reason>",
          "extracted_text": "<first 50 words from the original clause in the contract>"
        }},
        ...
      ]
    }}

    Return ONLY the JSON object. Do not include any explanation outside the JSON.
    """

# ---------- LLM Router ---------- #
def analyze_contract(text: str, model: str, clause_ids: list[int]) -> dict:
    # Load clauses
    clauses = load_selected_clauses(clause_ids)
    if not isinstance(clauses, list) or not all(isinstance(c, dict) for c in clauses):
        raise HTTPException(status_code=500, detail="Invalid clause format from DB")

    # Init embeddings + vector store
    embedding_model = HuggingFaceEmbeddings(model_name="sentence-transformers/all-MiniLM-L6-v2")
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=200)
    chunks = text_splitter.split_text(text)
    vector_store = FAISS.from_texts(chunks, embedding_model)
    retriever = vector_store.as_retriever(search_kwargs={"k": 3})

    # Step 1: Generate summary
    summary_query = "Contract overview: parties involved, main purpose, key obligations"
    summary_docs = retriever.invoke(summary_query)
    summary_context = "\n\n".join([d.page_content for d in summary_docs])
    summary_prompt = build_summary_prompt(summary_context)

    # Call LLM for summary
    summary_result = None
    if model == "1":
        llm = Ollama(model="mistral", temperature=0)
        summary_result = llm.invoke(summary_prompt).strip()
    elif model == "2":
        gemini_model = genai.GenerativeModel("models/gemini-2.5-pro")
        summary_result = gemini_model.generate_content(summary_prompt).text.strip()
    elif model == "3":
        response = query_endpoint(summary_prompt)
        parsed = json.loads(response)
        summary_result = parsed["response"] if "response" in parsed else parsed
    else:
        raise HTTPException(status_code=400, detail="Unsupported model.")

    # Step 2: Analyze clauses (existing logic)
    final_results = []
    batch_size = 3
    for i in range(0, len(clauses), batch_size):
        batch = clauses[i:i + batch_size]
        contexts = {}
        for clause in batch:
            query = f"{clause['clause_name']}: {clause['clause_text']}"
            docs = retriever.invoke(query)
            context = "\n\n".join([d.page_content for d in docs])
            contexts[clause['clause_name']] = context

        clauses_section = "\n\n".join(
            [f"Clause Name: {c['clause_name']}\nRequired Clause Text: {c['clause_text']}\nRelevant Contract Excerpts: {contexts.get(c['clause_name'], '')}" for c in batch]
        )
        prompt = build_prompt(summary_context, clauses_section, model)

        # Call LLM for clause analysis
        raw_result = None
        if model == "1":
            llm = Ollama(model="mistral", temperature=0)
            raw_result = llm.invoke(prompt).strip()
        elif model == "2":
            gemini_model = genai.GenerativeModel("models/gemini-2.5-pro")
            raw_result = gemini_model.generate_content(prompt).text.strip()
        elif model == "3":
            response = query_endpoint(prompt)
            parsed = json.loads(response)
            raw_result = parsed["response"] if "response" in parsed else parsed
        else:
            raise HTTPException(status_code=400, detail="Unsupported model.")

        # Parse JSON safely
        if isinstance(raw_result, str):
            cleaned_result = raw_result.replace("```json", "").replace("```", "").strip()
            if not cleaned_result:
                raise HTTPException(status_code=500, detail="Empty response from LLM")
            try:
                parsed_batch = safe_json_loads(raw_result if isinstance(raw_result, str) else json.dumps(raw_result))
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Invalid JSON from LLM: {str(e)}")

        elif isinstance(raw_result, dict):
            parsed_batch = raw_result
        else:
            raise HTTPException(status_code=500, detail="Unexpected result type.")

        # Merge analysis results
        if "analysis" in parsed_batch:
            final_results.extend(parsed_batch["analysis"])

    # Step 3: Combine summary and analysis
    return {
        "description": summary_result or "Summary could not be generated.",
        "analysis": final_results
    }

client = MongoClient("mongodb://localhost:27017/")
db = client["Contract-Compliance-Analyzer"]
contracts_cache = db["contracts_cache"]

def compute_cache_key(file_bytes: bytes, clause_ids: list[int], user_id: str) -> str:
    normalized_clauses = sorted(clause_ids)
    combined = {
        "pdf_hash": hashlib.sha256(file_bytes).hexdigest(),
        "clauses": normalized_clauses,
        "user_id": user_id
    }
    return hashlib.sha256(json.dumps(combined, sort_keys=True).encode()).hexdigest()

from datetime import datetime
import os

UPLOAD_DIR = Path(__file__).resolve().parent / "uploaded_contracts"
UPLOAD_DIR.mkdir(exist_ok=True)  # make sure base folder exists

@app.post("/analyze-contract", tags=["Contract Compliance"])
async def analyze_contract_api(
    file: UploadFile = File(...),
    model: str = Form(...),
    clauses: str = Form(...),
    user_id: str = Form(...),
    country_id: str = Form(...),
    domain_id: str = Form(...)
    ):
    start_time = datetime.utcnow()
    analysis_doc_id = None

    try:
        clause_ids_list = sorted([int(c.strip()) for c in clauses.strip("[]").split(",") if c.strip()])
        file_bytes = await file.read()

        # ✅ Compute cache key first
        pdf_hash = compute_cache_key(file_bytes, clause_ids_list, user_id)

        # ✅ Save uploaded PDF in folder
        contract_dir = UPLOAD_DIR / pdf_hash
        contract_dir.mkdir(parents=True, exist_ok=True)

        pdf_path = contract_dir / file.filename
        with open(pdf_path, "wb") as f:
            f.write(file_bytes)

        # ✅ Check cache first (no new log if cached result is valid)
        cached_doc = contracts_cache.find_one({"_id": pdf_hash, "user_id": user_id})
        if cached_doc and cached_doc.get("feedback", "").lower() != "dislike":
            return {
                "analysis": cached_doc.get("analysis", []),
                "cachekey": pdf_hash,
                "cached": True,
                "saved_pdf": str(pdf_path)
            }

        # ✅ Insert initial log
        analysis_doc_id = user_analysis.insert_one({
            "user_id": user_id,
            "cached_response_id": pdf_hash,
            "pdfName": file.filename,
            "start_time": start_time,
            "end_time": None,
            "is_success": False
        }).inserted_id

        # Extract text and analyze
        text = extract_text_from_pdf(file_bytes)
        result = analyze_contract(text, model, clause_ids_list)

        # Save new record
        contracts_cache.update_one(
            {"_id": pdf_hash},
            {
                "$set": {
                    "user_id": user_id,
                    "pdf_name": file.filename,
                    "analysis": result,
                    "clauses": clause_ids_list,
                    "feedback": "",
                    "feedback_given": False,
                    "uploaded_at": datetime.utcnow().isoformat(),
                    "country_id": country_id,
                    "domain_id": domain_id
                }
            },
            upsert=True
        )

        user_analysis.update_one(
            {"_id": analysis_doc_id},
            {"$set": {"end_time": datetime.utcnow(), "is_success": True}}
        )

        return {
            "analysis": result,
            "cachekey": pdf_hash,
            "cached": False,
            "saved_pdf": str(pdf_path)  # ✅ return saved path
        }

    except Exception as e:
        if analysis_doc_id:
            user_analysis.update_one(
                {"_id": analysis_doc_id},
                {"$set": {"end_time": datetime.utcnow(), "is_success": False}}
            )

        return {
            "error": str(e),
            "cached": False
        }


# #-----------Compute PDF Hash--------#
def compute_pdf_hash(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()

#---------Read Cache File if Exists-----#
CACHE_DIR = Path(__file__).resolve().parent / "json_cache"
CACHE_DIR.mkdir(exist_ok=True)

def load_cached_response(pdf_hash):
    cache_file = f"./cache/{pdf_hash}.json"
    if os.path.exists(cache_file):
        with open(cache_file, "r", encoding="utf-8") as f:  # "r" mode for reading
            return json.load(f)
    return None

def save_response_to_cache(pdf_hash: str, data: dict):
    cache_path = CACHE_DIR / f"{pdf_hash}.json"
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

from pydantic import BaseModel
from fastapi import HTTPException

class FeedbackInput(BaseModel):
    feedback: str
    cache_key: str   # _id of the document

@app.post("/feedback", tags=["Contract Compliance"])
def submit_feedback(feedback_input: FeedbackInput):
    # Find the contract by cache_key
    contract = contracts_cache.find_one({"_id": feedback_input.cache_key})
    if not contract:
        raise HTTPException(status_code=404, detail="Cache key not found")

    # Update the feedback and set feedback_given = True
    contracts_cache.update_one(
        {"_id": feedback_input.cache_key},
        {"$set": {
            "feedback": feedback_input.feedback,
            "feedback_given": True
        }}
    )

    return {"message": "Feedback submitted successfully"}

import json
import tempfile
import shutil

def atomic_write_json(data, filepath):
    with tempfile.NamedTemporaryFile("w", delete=False, dir=os.path.dirname(filepath)) as tmpfile:
        json.dump(data, tmpfile, indent=2)
        tempname = tmpfile.name
    shutil.move(tempname, filepath)

import hashlib

def generate_cache_key(text: str, country: str) -> str:
    """
    Generate a unique cache key based on PDF text and country.
    """
    combined = f"{text}-{country}"
    return hashlib.sha256(combined.encode("utf-8")).hexdigest()

from fastapi import FastAPI
from pydantic import BaseModel
from bson import ObjectId


class Country(BaseModel):
    country_name: str

class Domain(BaseModel):
    domain_name: str
    country_id: int

class Compliance(BaseModel):
    compliance_name: str
    domain_id: int

class Clause(BaseModel):
    clause_name:str
    clause_text: str
    domain_id: Optional[int] = None

# ---------- Auto Increment Function ----------
def get_next_sequence(name):
    counter = db.counters.find_one_and_update(
        {"_id": name},
        {"$inc": {"sequence_value": 1}},
        upsert=True,
        return_document=True
    )
    return counter["sequence_value"]


from fastapi import HTTPException

# ---------- CRUD for Country ----------
@app.post("/countries",tags=["Countries"])
def create_country(country: Country):
    next_id = get_next_sequence("country_id")
    db.countries.insert_one({"_id": next_id, "country_name": country.country_name})
    return {"message": "Country added", "id": next_id}

@app.get("/countries",tags=["Countries"])
def get_countries():
    return list(db.countries.find({}, {"_id": 1, "country_name": 1}))

@app.put("/countries/{country_id}",tags=["Countries"])
def update_country(country_id: int, country: Country):
    result = db.countries.update_one({"_id": country_id}, {"$set": {"country_name": country.country_name}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Country not found")
    return {"message": "Country updated"}

@app.delete("/countries/{country_id}",tags=["Countries"])
def delete_country(country_id: int):
    result = db.countries.delete_one({"_id": country_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Country not found")
    return {"message": "Country deleted"}

@app.get("/countries/{country_id}", tags=["Countries"])
def get_country_by_id(country_id: int):
    country = db.countries.find_one({"_id": country_id}, {"_id": 1, "country_name": 1})
    if not country:
        raise HTTPException(status_code=404, detail="Country not found")
    return country


# ===== DOMAIN CRUD =====
@app.post("/domains",tags=["Domains"])
def create_domain(domain: Domain):
    next_id = get_next_sequence("domain_id")
    db.domains.insert_one({
        "_id": next_id,
        "domain_name": domain.domain_name,
        "country_id": domain.country_id
    })
    return {"message": "Domain added", "id": next_id}

@app.get("/domains", tags=["Domains"])
def get_domains():
    pipeline = [
        {
            "$lookup": {
                "from": "countries",        # Join with countries collection
                "localField": "country_id", # Field in domains
                "foreignField": "_id",      # Field in countries
                "as": "country_info"
            }
        },
        {"$unwind": "$country_info"},       # Flatten the array
        {
            "$project": {                    # Select fields to return
                "_id": 1,
                "domain_name": 1,
                "country_id": 1,
                "country_name": "$country_info.country_name"
            }
        }
    ]
    
    domains = list(db.domains.aggregate(pipeline))
    return domains 


@app.get("/domains/byid/{domain_id}", tags=["Domains"])
def get_domain_by_id(domain_id: int):
    pipeline = [
        {"$match": {"_id": domain_id}},
        {
            "$lookup": {
                "from": "countries",
                "localField": "country_id",
                "foreignField": "_id",
                "as": "country_info"
            }
        },
        {"$unwind": "$country_info"},
        {
            "$project": {
                "_id": 1,
                "domain_name": 1,
                "country_id": 1,
                "country_name": "$country_info.country_name"
            }
        }
    ]
    
    result = list(db.domains.aggregate(pipeline))
    if not result:
        raise HTTPException(status_code=404, detail="Domain not found")
    return result[0]

@app.put("/domains/{domain_id}",tags=["Domains"])
def update_domain(domain_id: int, domain: Domain):
    result = db.domains.update_one({"_id": domain_id}, {"$set": {
        "domain_name": domain.domain_name,
        "country_id": domain.country_id
    }})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Domain not found")
    return {"message": "Domain updated"}

@app.delete("/domains/{domain_id}",tags=["Domains"])
def delete_domain(domain_id: int):
    result = db.domains.delete_one({"_id": domain_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Domain not found")
    return {"message": "Domain deleted"}

@app.get("/domains/bycountry/{country_id}", tags=["Domains"])
def get_domains_by_country(country_id: int):
    pipeline = [
        {"$match": {"country_id": country_id}},
        {
            "$lookup": {
                "from": "countries",        # Join with countries collection
                "localField": "country_id", # Field in domains
                "foreignField": "_id",      # Field in countries
                "as": "country_info"
            }
        },
        {"$unwind": "$country_info"},       # Flatten array
        {
            "$project": {                    # Fields to return
                "_id": 1,
                "domain_name": 1,
                "country_id": 1,
                "country_name": "$country_info.country_name"
            }
        }
    ]

    domains = list(db.domains.aggregate(pipeline))
    return domains

# ===== COMPLIANCE CRUD =====
@app.get("/compliances", tags=["Compliances"])
def get_all_compliances():
    pipeline = [
        {
            "$lookup": {
                "from": "domains",        # Join with domains collection
                "localField": "domain_id", # Field in compliances
                "foreignField": "_id",     # Field in domains
                "as": "domain_info"
            }
        },
        {"$unwind": "$domain_info"},    # Flatten array to object
        {
            "$project": {               # Select fields to return
                "_id": 1,
                "compliance_name": 1,
                "domain_id": 1,
                "domain_name": "$domain_info.domain_name"
            }
        }
    ]
    
    compliances = list(db.compliances.aggregate(pipeline))
    return compliances

@app.post("/compliances",tags=["Compliances"])
def create_compliance(compliance: Compliance):
    next_id = get_next_sequence("compliance_id")
    db.compliances.insert_one({
        "_id": next_id,
        "compliance_name": compliance.compliance_name,
        "domain_id": compliance.domain_id
    })
    return {"message": "Compliance added", "id": next_id}

@app.get("/compliances/{domain_id}", tags=["Compliances"])
def get_compliances(domain_id: int):
    pipeline = [
        {"$match": {"domain_id": domain_id}},  # Filter by domain_id
        {
            "$lookup": {
                "from": "domains",          # Join with domains collection
                "localField": "domain_id",  # Field in compliances
                "foreignField": "_id",      # Field in domains
                "as": "domain_info"
            }
        },
        {"$unwind": "$domain_info"},            # Flatten array
        {
            "$project": {
                "_id": 1,
                "compliance_name": 1,
                "domain_id": 1,
                "domain_name": "$domain_info.domain_name"
            }
        }
    ]

    compliances = list(db.compliances.aggregate(pipeline))
    return compliances

@app.put("/compliances/{compliance_id}",tags=["Compliances"])
def update_compliance(compliance_id: int, compliance: Compliance):
    result = db.compliances.update_one({"_id": compliance_id}, {"$set": {
        "compliance_name": compliance.compliance_name,
        "domain_id": compliance.domain_id
    }})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Compliance not found")
    return {"message": "Compliance updated"}

@app.delete("/compliances/{compliance_id}",tags=["Compliances"])
def delete_compliance(compliance_id: int):
    result = db.compliances.delete_one({"_id": compliance_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Compliance not found")
    return {"message": "Compliance deleted"}

# ===== CLAUSE CRUD =====
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from typing import Optional

class Clause(BaseModel):
    clause_name: str
    clause_text: str
    domain_id: Optional[int] = None

# ===== CLAUSE CRUD =====

# Create Clause
@app.post("/clauses", tags=["Clauses"])
def create_clause(clause: Clause):
    next_id = get_next_sequence("clause_id")  # Your ID generator
    db.clauses.insert_one({
        "_id": next_id,
        "clause_name": clause.clause_name,
        "clause_text": clause.clause_text,
        "domain_id": clause.domain_id
    })
    return {"message": "Clause added", "id": next_id}

# Get all clauses with domain info
@app.get("/clauses", tags=["Clauses"])
def get_all_clauses():
    pipeline = [
        {
            "$lookup": {
                "from": "domains",
                "localField": "domain_id",
                "foreignField": "_id",
                "as": "domain_info"
            }
        },
        {"$unwind": {"path": "$domain_info", "preserveNullAndEmptyArrays": True}},
        {
            "$lookup": {
                "from": "countries",
                "localField": "domain_info.country_id",
                "foreignField": "_id",
                "as": "country_info"
            }
        },
        {"$unwind": {"path": "$country_info", "preserveNullAndEmptyArrays": True}},
        {
            "$project": {
                "_id": 1,
                "clause_name": 1,
                "clause_text": 1,
                "domain_id": 1,
                "domain_name": "$domain_info.domain_name",
                "country_id": "$country_info._id",
                "country_name": "$country_info.country_name"
            }
        }
    ]

    clauses = list(db.clauses.aggregate(pipeline))
    return clauses

# Get clauses by country & domain (filter)
@app.get("/clauses/filter", tags=["Clauses"])
def get_clauses_by_country_domain(
    country_id: int = Query(..., description="Country ID"),
    domain_id: int = Query(..., description="Domain ID")
    ):
    pipeline = [
        {"$match": {"domain_id": domain_id}},
        {
            "$lookup": {
                "from": "domains",
                "localField": "domain_id",
                "foreignField": "_id",
                "as": "domain_info"
            }
        },
        {"$unwind": "$domain_info"},
        {"$match": {"domain_info.country_id": country_id}},
        {
            "$project": {
                "_id": 1,
                "clause_name": 1,
                "clause_text": 1,
                "domain_id": 1,
                "domain_name": "$domain_info.domain_name",
                "country_id": "$domain_info.country_id"
            }
        }
    ]
    clauses = list(db.clauses.aggregate(pipeline))
    return clauses

# Get clauses by domain (single path parameter)
@app.get("/clauses/by-domain/{domain_id}", tags=["Clauses"])
def get_clauses_by_domain(domain_id: int):
    clauses = list(db.clauses.find(
        {"domain_id": domain_id},
        {"_id": 1, "clause_name": 1, "clause_text": 1, "domain_id": 1}
    ))
    return clauses

# Update Clause
@app.put("/clauses/{clause_id}", tags=["Clauses"])
def update_clause(clause_id: int, clause: Clause):
    result = db.clauses.update_one({"_id": clause_id}, {"$set": {
        "clause_name": clause.clause_name,
        "clause_text": clause.clause_text,
        "domain_id": clause.domain_id
    }})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Clause not found")
    return {"message": "Clause updated"}

# Get Clause by ID
@app.get("/clauses/{clause_id}", tags=["Clauses"])
def get_clause_by_id(clause_id: int):
    clause = db.clauses.find_one(
        {"_id": clause_id},
        {"_id": 1, "clause_name": 1, "clause_text": 1, "domain_id": 1}
    )

    if not clause:
        raise HTTPException(status_code=404, detail="Clause not found")

    # Fetch domain details
    domain = db.domains.find_one(
        {"_id": clause["domain_id"]},
        {"_id": 1, "domain_name": 1}
    )

    # Add domain_name if found
    clause["domain_name"] = domain["domain_name"] if domain else None

    return clause

# Delete Clause
@app.delete("/clauses/{clause_id}", tags=["Clauses"])
def delete_clause(clause_id: int):
    result = db.clauses.delete_one({"_id": clause_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Clause not found")
    return {"message": "Clause deleted"}


#--------------------------------------------chat Bot-------------------------------------------------
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_community.vectorstores import FAISS
from langchain_huggingface import HuggingFaceEmbeddings
from fastapi import UploadFile, File, Form, HTTPException
from pydantic import BaseModel
import json
import re
from ollama import Client
import google.generativeai as genai
import logging
from typing import Optional, List, Dict, Any
from pymongo import MongoClient
from bson import ObjectId

# Set up logging
logging.basicConfig(level=logging.WARNING)
logger = logging.getLogger(__name__)

# Embedding model
embedding_model = HuggingFaceEmbeddings(model_name="sentence-transformers/all-MiniLM-L6-v2")

# In-memory FAISS cache
vector_store_cache = {}

class ChatbotFileInput(BaseModel):
    question: str
    model: str  # "1" = Mistral, "2" = Gemini

class ChatbotMetadataInput(BaseModel):
    question: str
    model: str

import re

def normalize_query(q: str) -> str:
    """
    Normalize synonyms in the query so matching works consistently.
    """
    replacements = {
        # Domain synonyms
        r"\bindustry\b": "domain",
        r"\bindustries\b": "domains",
        r"\bsector\b": "domain",
        r"\bsectors\b": "domains",

        # Country synonyms
        r"\bnation\b": "country",
        r"\bnations\b": "countries",
        r"\bstates?\b": "countries",
        r"\bregion\b": "country",
        r"\bregions\b": "countries",

        # Clause synonyms
        r"\bprovision\b": "clause",
        r"\bprovisions\b": "clauses",
        r"\barticle\b": "clause",
        r"\barticles\b": "clauses",
        r"\bterm\b": "clause",
        r"\bterms\b": "clauses",
    }

    for pattern, repl in replacements.items():
        q = re.sub(pattern, repl, q, flags=re.IGNORECASE)

    return q.lower().strip()

def clean_entity_name(name: str) -> str:
    """
    Remove trailing keywords like domain, industry, clause, etc.
    Example: "oil and gas domain" -> "oil and gas"
    """
    stopwords = [
        "domain", "domains", "industry", "industries",
        "sector", "sectors", "country", "countries",
        "clause", "clauses", "provision", "provisions",
        "article", "articles", "term", "terms"
    ]
    tokens = name.split()
    if tokens and tokens[-1].lower() in stopwords:
        tokens = tokens[:-1]
    return " ".join(tokens).strip()

def is_greeting(question: str) -> bool:
    """Check if the question is a greeting"""
    greetings = [
        r'\b(hi|hello|hey|good morning|good afternoon|good evening)\b',
        r'^\s*(hi|hello|hey)\s*[!.?]*\s*$'
    ]
    question_lower = question.lower().strip()
    return any(re.search(pattern, question_lower, re.IGNORECASE) for pattern in greetings)

def is_contract_related(question: str) -> bool:
    q = question.lower().strip()

    # Strong contract indicators
    contract_phrases = [
        "in the contract", "does the contract have", "is there any",
        "under this contract", "mentioned in the contract"
    ]

    # Contract-related keywords (weaker signals)
    contract_keywords = [
        "termination", "payment", "confidentiality", "governing law",
        "liability", "obligation", "condition", "agreement", "sub-clause",
        "section", "rights", "responsibilities"
    ]

    # 1. If any strong phrase is present → definitely contract
    if any(phrase in q for phrase in contract_phrases):
        return True

    # 2. If keywords appear *with contract context words*
    if any(kw in q for kw in contract_keywords):
        if "clause" in q or "contract" in q or "agreement" in q:
            return True

    return False

def is_metadata_query(question: str) -> bool:
    """Check if the question is related to metadata (countries, domains, clauses)"""
    question = question.lower().strip()
    metadata_intents = [
        "list", "show", "give me", "describe", "summary", "summarize",
        "details of", "what are", "explain", "description of", "information on"
    ]
    metadata_keywords = [
        'country', 'countries', 'domain', 'domains', 'clause', 'clauses',
        'what countries', 'which countries', 'list countries', 'show countries',
        'what domains', 'which domains', 'list domains', 'show domains',
        'what clauses', 'which clauses', 'list clauses', 'show clauses',
        'available countries', 'available domains', 'available clauses',
        'force majeure', 'liability', 'oil and gas', 'metadata'
    ]

    if any(intent in question for intent in metadata_intents):
        if any(word in question for word in metadata_keywords):
            return True

    # General metadata lookups
    if "metadata" in question.lower():
        return True

    return False
    
def detect_intent(question: str) -> str:
    """Detect whether the user wants list, count, existence, or detail"""
    q = question.lower()

    if any(kw in q for kw in ["how many", "number of", "count of"]):
        return "count"
    if any(kw in q for kw in ["are there", "is there", "does", "do "]):
        return "existence"
    if any(kw in q for kw in ["list", "show", "which", "available", "what"]):
        return "list"
    if any(kw in q for kw in ["explain", "describe", "summarize", "detail", "tell me about", "give me", "description", "summary", "information on"]):
        return "detail"
    return "unknown"


def search_metadata_entities(question: str) -> tuple:
    """
    Return (countries, domains, clauses) based on entity keywords in question,
    with robust joins (countries → domains → clauses).
    """
    q = normalize_query(question)   # normalize synonyms first

    # --- MULTI-JOIN ---
    # Example: "how many clauses are there for USA in oil and gas"
    if "clause" in q and ("for" in q and "in" in q):
        match = re.search(r'for (.+?) in (.+)', q)
        if match:
            country_name = clean_entity_name(match.group(1).strip())
            domain_name = clean_entity_name(match.group(2).strip())

            pipeline = [
                {"$lookup": {
                    "from": "domains",
                    "localField": "domain_id",
                    "foreignField": "_id",
                    "as": "domain_info"
                }},
                {"$unwind": "$domain_info"},
                {"$lookup": {
                    "from": "countries",
                    "localField": "domain_info.country_id",
                    "foreignField": "_id",
                    "as": "country_info"
                }},
                {"$unwind": "$country_info"},
                {"$match": {
                    "domain_info.domain_name": {"$regex": domain_name, "$options": "i"},
                    "country_info.country_name": {"$regex": country_name, "$options": "i"}
                }}
            ]
            clauses = list(db.clauses.aggregate(pipeline))
            return [], [], clauses

    # --- SINGLE-JOIN ---
    # Countries for a domain
    if "countries" in q and "for" in q:
        domain_match = re.search(r'for (.+)', q)
        if domain_match:
            domain_name = clean_entity_name(domain_match.group(1).strip())
            pipeline = [
                {"$match": {"domain_name": {"$regex": domain_name, "$options": "i"}}},
                {"$lookup": {
                    "from": "countries",
                    "localField": "country_id",
                    "foreignField": "_id",
                    "as": "country_info"
                }},
                {"$unwind": "$country_info"}
            ]
            domains = list(db.domains.aggregate(pipeline))
            countries = [d["country_info"] for d in domains if "country_info" in d]
            return countries, [], []

    # Clauses for a domain
    if "clauses" in q and "for" in q:
        domain_match = re.search(r'for (.+)', q)
        if domain_match:
            domain_name = clean_entity_name(domain_match.group(1).strip())
            pipeline = [
                {"$lookup": {
                    "from": "domains",
                    "localField": "domain_id",
                    "foreignField": "_id",
                    "as": "domain_info"
                }},
                {"$unwind": "$domain_info"},
                {"$match": {"domain_info.domain_name": {"$regex": domain_name, "$options": "i"}}}
            ]
            clauses = list(db.clauses.aggregate(pipeline))
            return [], [], clauses

    # Domains in a country
    if "domains" in q and "in" in q:
        country_match = re.search(r'in (.+)', q)
        if country_match:
            country_name = clean_entity_name(country_match.group(1).strip())
            country = db.countries.find_one({
                "country_name": {"$regex": country_name, "$options": "i"}
            })
            if country:
                domains = list(db.domains.find({"country_id": country["_id"]}))
                return [], domains, []

    # --- DIRECT COLLECTION QUERIES ---
    if "country" in q or "countries" in q:
        return list(db.countries.find({})), [], []
    if "domain" in q or "domains" in q:
        return [], list(db.domains.find({})), []
    if "clause" in q or "clauses" in q:
        return [], [], list(db.clauses.find({}))

    return [], [], []

def format_metadata_response(question: str, intent: str, countries: List[Dict], domains: List[Dict], clauses: List[Dict]) -> str:
    """Format metadata results based on intent classification, with join handling"""

    # --- COUNT intent ---
    if intent == "count":
        print("Count intent detected")
        if "country" in question.lower() or "countries" in question.lower():
            print("Counting countries")
            if countries:
                print("Countries found:", countries)
                return f"There {'is' if len(countries)==1 else 'are'} {len(countries)} country{'s' if len(countries)>1 else ''}: " + ", ".join(c.get("country_name","Unknown") for c in countries)
            return "There are 0 countries."
        if "domain" in question.lower() or "domains" in question.lower():
            print("Counting domains")
            if domains:
                print("Domains found:", domains)
                return f"There {'is' if len(domains)==1 else 'are'} {len(domains)} domain{'s' if len(domains)>1 else ''}: " + ", ".join(d.get("domain_name","Unknown") for d in domains)
            return "There are 0 domains."
        if "clause" in question.lower() or "clauses" in question.lower():
            if clauses:
                return f"There {'is' if len(clauses)==1 else 'are'} {len(clauses)} clause{'s' if len(clauses)>1 else ''}: " + ", ".join(c.get("clause_name","Unknown") for c in clauses)
            return "There are 0 clauses."
        return "No matching metadata found to count."

        # --- EXISTENCE intent ---
    if intent == "existence":
        if countries: return f"There {'is' if len(countries)==1 else 'are'} {len(countries)} matching countr{'y' if len(countries)==1 else 'ies'}."
        if domains: return f"There {'is' if len(domains)==1 else 'are'} {len(domains)} matching domain{'s' if len(domains)!=1 else ''}."
        if clauses: return f"There {'is' if len(clauses)==1 else 'are'} {len(clauses)} matching clause{'s' if len(clauses)!=1 else ''}."
        return "No matching results found."

    # --- DETAIL intent ---
    if intent == "detail" and clauses:
        q = question.lower()
        filtered_clauses = [
            c for c in clauses 
            if any(kw in c.get("clause_name", "").lower() for kw in q.split())
        ]

        # If no exact match, fallback to all
        target_clauses = filtered_clauses if filtered_clauses else clauses

        details = []
        for c in target_clauses:
            details.append({
                "clause_name": c.get("clause_name", "Unknown"),
                # "domain": (c.get("domain_info") or [{}])[0].get("domain_name", "Unknown"),
                # "country": (c.get("country_info") or [{}])[0].get("country_name", "Unknown"),
                "clause_summary": c.get("clause_text", "")[:300] + "..."
            })

        #details = []
        #for c in clauses:
        #    details.append({
        #        "name": c.get("clause_name", "Unknown"),
        #        "domain": (c.get("domain_info") or [{}])[0].get("domain_name", "Unknown"),
        #        "country": (c.get("country_info") or [{}])[0].get("country_name", "Unknown"),
        #        "text": c.get("clause_text", "")[:300] + "..."
        #    })
        return json.dumps(details, indent=2)

    # --- LIST intent (fallback if no details) ---
    if intent in ["list", "unknown"]:
        if countries: return "Countries: " + ", ".join(c.get("country_name","Unknown") for c in countries)
        if domains: return "Domains: " + ", ".join(d.get("domain_name","Unknown") for d in domains)
        if clauses: return "Clauses: " + ", ".join(c.get("clause_name","Unknown") for c in clauses)

    return "No metadata results found."

import re

def is_valid_answer(answer: str) -> bool:
    if not answer or not answer.strip():
        return False

    # Whitelisted safe fallback responses
    safe_fallbacks = [
        "sorry, i can't answer this query based on the contract or metadata"
    ]
    if answer.strip().lower() in safe_fallbacks:
        return True

    # General error-like patterns
    error_patterns = [
        r'(^|\n)\s*error',
        r'\bsorry[, ]',             
        r'couldn\'t process',       
        r'can[\' ]?t answer',       
        r'not able to',             
        r'failed to',               
        r'exception',               
        r'invalid',                 
        r'json.*error',             
    ]

    ans_lower = answer.lower()
    for pattern in error_patterns:
        if re.search(pattern, ans_lower, flags=re.IGNORECASE):
            return False

    return True


# Update your chatbot_metadata function to use the enhanced search:
@app.post("/chatbot-metadata", tags=["Contract Chatbot"])
async def chatbot_metadata(input_data: ChatbotMetadataInput):
    try:
        question = input_data.question
        model = input_data.model

        # Greeting
        if is_greeting(question):
            return {
                "answer": "Hi, I am Contract Genie. I can help you with information about countries, domains, and clauses in our database. What would you like to know?",
                "cachekey": "greeting_metadata"
            }

        # Cache check
        cache_key = f"metadata_{question}_{model}"
        #cached = db.contracts_cache.find_one({"_id": cache_key})
        #if cached and cached.get("feedback", "").lower() != "dislike":
        #    return {"answer": cached["answer"], "cachekey": "metadata"}
        cached_response = db.contracts_cache.find_one({"_id": cache_key})
        if cached_response and cached_response.get("answer"):
            return {"answer": cached_response["answer"], "cachekey": "metadata"}

        # Detect intent + search
        intent = detect_intent(question)
        print(f"Detected intent: {intent}")
        countries, domains, clauses = search_metadata_entities(question)
        print(f"Found {countries} countries, {domains} domains, {clauses} clauses")

        # Format response
        answer = format_metadata_response(question, intent, countries, domains, clauses)

        # Cache
        if is_valid_answer(answer):
            db.contracts_cache.update_one(
                {"_id": cache_key},
                {"$set": {
                    "answer": answer,
                    "model": model,
                    "question": question,
                    "feedback": "",
                    "feedback_given": False
                }},
                upsert=True
            )

        return {"answer": answer, "cachekey": "metadata"}

    except Exception as e:
        logger.error(f"Metadata chatbot error: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Metadata chatbot error: {str(e)}")
    
#def is_valid_answer(answer: str) -> bool:
#    if not answer:
#        return False
#    bad_signals = [
#        "error while processing query",
#        "i'm getting error",
#        "Sorry, I couldn't process the contract content.",
#        "Sorry I can't answer this query",
#        "llm did not return valid json"
#    ]
#    return not any(bad in answer.lower() for bad in bad_signals)

import requests
#-------------Endpoint for LLM Model----------------
#def query_endpoint(param_string):
#    url = "http://172.16.117.136:8000/query"
#    # Preserve JSON structure, only trim outer whitespace
#    cleaned_string = param_string.strip() if param_string else ""
#    payload = {"prompt": cleaned_string}
#    try:
#        print("Payload:", payload)
#        response = requests.post(url, json=payload)
#        print("Status code:", response.status_code)
#        print("Raw response:", response.text)
#        response.raise_for_status()
#        return response.text
#    except requests.RequestException as e:
#        print("Request error:", str(e))
#        return f"Error: {str(e)}"

#------------- Using Ollama for local LLM ----------------
import subprocess
def query_endpoint(param_string: str):
    cleaned_string = param_string.strip() if param_string else ""
    try:
        print("Payload:", cleaned_string)
        
        # Call local llama3.1 using Ollama
        result = subprocess.run(
            ["ollama", "run", "llama3.1"],
            input=cleaned_string,   # pass string, not bytes
            capture_output=True,
            text=True
        )
        
        if result.returncode != 0:
            raise Exception(result.stderr)
        
        answer = result.stdout.strip()
        print("LLM response:", answer)
        return answer

    except Exception as e:
        print("LLM error:", str(e))
        return f"Error: {str(e)}"

from fastapi import Form, File, UploadFile, HTTPException
@app.post("/chatbot-file", tags=["Contract Chatbot"])
async def chatbot_file(
    file: UploadFile = File(...),
    question: str = Form(...),
    model: str = Form(...)
):
    try:
        # Step 0: Greetings
        if is_greeting(question):
            return {
                "answer": "Hi, I am Contract Genie. I can help you with metadata queries (countries, domains, clauses) or your uploaded contract.",
                "cachekey": "greeting"
            }

        # Step 1: Read file + hash
        file_bytes = await file.read()
        pdf_hash = compute_pdf_hash(file_bytes)

        # Step 2: Cache check
        cache_key = f"{pdf_hash}_{question}_{model}"
        cached_response = db.contracts_cache.find_one({"_id": cache_key})
        if cached_response and cached_response.get("answer"):
            return {"answer": cached_response["answer"], "cachekey": pdf_hash}

        # Step 3: Query type classification
        #is_contract_q = is_contract_related(question)   # stronger check
        is_metadata_q = is_metadata_query(question)
        is_contract_q = is_contract_related(question)
        answer = ""
        
        if is_metadata_q:
            intent = detect_intent(question)
            countries, domains, clauses = search_metadata_entities(question)
            answer = format_metadata_response(question, intent, countries, domains, clauses)

        # --- Contract-related queries ---
        elif is_contract_q:
            # Build FAISS index if not cached
            if pdf_hash not in vector_store_cache:
                text = extract_text_from_pdf(file_bytes)
                if text.strip():
                    text_splitter = RecursiveCharacterTextSplitter(chunk_size=500, chunk_overlap=100)
                    chunks = text_splitter.split_text(text)
                    vector_store = FAISS.from_texts(chunks, embedding_model)
                    vector_store_cache[pdf_hash] = vector_store
            vector_store = vector_store_cache.get(pdf_hash)

            context = ""
            if vector_store:
                retriever = vector_store.as_retriever(search_kwargs={"k": 2})
                docs = retriever.invoke(question)
                context = "\n\n".join([d.page_content for d in docs])

            if context.strip():
                prompt = f"""
                You are an expert contract assistant. Answer the question based ONLY on the provided excerpts.

                Context:
                {context}

                Question:
                {question}

                Instructions:
                - Be concise, plain text.
                - If not found in the context, reply: "Sorry I can't answer this query, try another query."
                """
                try:
                    raw_response = query_endpoint(prompt)
                    # If LLM returns plain text, just use it
                    if isinstance(raw_response, str):
                        answer = raw_response.strip()
                    elif isinstance(raw_response, dict) and "response" in raw_response:
                        answer = raw_response["response"]
                    else:
                        answer = str(raw_response)
                except:
                    answer = "Sorry, I couldn't process the contract content."
        #elif is_metadata_q:
        #    intent = detect_intent(question)
        #    countries, domains, clauses = search_metadata_entities(question)
        #    answer = format_metadata_response(question, intent, countries, domains, clauses)
        
        # --- Fallback ---
        else:
            answer = "Sorry, I can't answer this query based on the contract or metadata."

        # Step 4: Cache
        if is_valid_answer(answer):
            db.contracts_cache.update_one(
                {"_id": cache_key},
                {"$set": {"answer": answer, "model": model, "question": question}},
                upsert=True
            )

        return {"answer": answer, "cachekey": pdf_hash}

    except Exception as e:
        logger.error(f"Chatbot error: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Chatbot error: {str(e)}")


import json
def create_or_overwrite_prompt_file(prompt_content: str, file_path: str = "prompt.txt") -> None:
    """
    Creates or overwrites a text file named prompt.txt with the given content as valid JSON.
    
    Args:
        prompt_content (str): The content to write to the file, expected to be a valid JSON string.
        file_path (str): The path to the file (default: 'prompt.txt').
    """
    try:
        # Validate that prompt_content is valid JSON
        json.loads(prompt_content)
        with open(file_path, "w", encoding="utf-8") as file:
            # Create JSON structure with prompt_content as the value of the "prompt" key
            json.dump({"prompt": prompt_content}, file, ensure_ascii=False, indent=2)
        print(f"Successfully created or overwrote {file_path} with valid JSON")
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON content: {str(e)}")
    except Exception as e:
        print(f"Error writing to {file_path}: {str(e)}")


#------------------------------------------------------------------------------------------------------------------------------------------------
import requests
import json
from fastapi import HTTPException

def query_endpoint(param_string):
    url = "http://172.16.117.136:8001/query"
    # Preserve JSON structure, only trim outer whitespace
    cleaned_string = param_string.strip() if param_string else ""
    payload = {"prompt": cleaned_string}
    try:
        print("Payload:", payload)
        response = requests.post(url, json=payload)
        print("Status code:", response.status_code)
        print("Raw response:", response.text)
        response.raise_for_status()
        return response.text
    except requests.RequestException as e:
        print("Request error:", str(e))
        return f"Error: {str(e)}"

def build_summary_prompt(context: str) -> str:
    """
    Build a prompt to generate a brief summary of the contract.
    """
    return f"""
    You are an expert contract reviewer. Based on the following contract excerpt, provide a brief summary of the contract in 2–3 sentences (at least 50 words). Include key parties, purpose, subject matter, and type of obligations.

    Contract Excerpt:
    {context}

    Instructions:
    - Return only the summary as plain text.
    - Ensure the summary is clear, concise, and meets the word count requirement.
    """

import re, json

def safe_json_loads(raw_result: str) -> dict:
    """
    Clean LLM output and parse as JSON safely.
    """
    if not raw_result:
        raise ValueError("Empty LLM response")

    # Remove code fences if any
    cleaned = raw_result.replace("```json", "").replace("```", "").strip()

    # Extract only the JSON part between the first { and the last }
    match = re.search(r"\{.*\}", cleaned, re.DOTALL)
    if not match:
        raise ValueError(f"No valid JSON object found in response: {cleaned[:200]}")
    
    json_str = match.group(0)

    return json.loads(json_str)



#---------------------Auth------------------------------------------------

from fastapi import Depends, HTTPException, status, FastAPI, File, UploadFile, Form, Query
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext
from datetime import datetime, timedelta
from pydantic import BaseModel, EmailStr
from pymongo.errors import DuplicateKeyError
from typing import Optional

# JWT Configuration
SECRET_KEY = "your-secret-key"  # Replace with a strong secret key in production
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# OAuth2 scheme
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/login")

# MongoDB collections
users_collection = db["users"]
roles_collection = db["roles"]

# Pydantic Models
class UserCreate(BaseModel):
    first_name: str
    last_name: str
    email: EmailStr
    password: str
    role_id: int = 2  # Default to 2 (User)

class UserInDB(BaseModel):
    first_name: str
    last_name: str
    email: EmailStr
    hashed_password: str
    role_id: int
    role_name: str

from pydantic import Field

class UserInfo(BaseModel):
    id: str = Field(..., alias="_id")
    first_name: str
    last_name: str
    email: str
    role_id: int
    role_name: str

    class Config:
        allow_population_by_field_name = True

class Token(BaseModel):
    access_token: str
    token_type: str
    user: UserInfo

class TokenData(BaseModel):
    email: Optional[str] = None
    role: Optional[str] = None

# Helper Functions
def get_role(role_id: int) -> Optional[dict]:
    """Retrieve a role from MongoDB by role_id."""
    return roles_collection.find_one({"_id": role_id})

def hash_password(password: str) -> str:
    """Hash a password using bcrypt."""
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    """Verify a plain password against a hashed password."""
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict) -> str:
    """Create a JWT access token."""
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def get_user(email: str) -> Optional[dict]:
    """Retrieve a user from MongoDB by email."""
    user = users_collection.find_one({"email": email})
    if user:
        role = get_role(user["role_id"])
        if role:
            user["role_name"] = role["role_name"]
        else:
            user["role_name"] = "Unknown"
    return user

async def get_current_user(token: str = Depends(oauth2_scheme)) -> dict:
    """Get the current user from the JWT token."""
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        role: str = payload.get("role")
        if email is None or role is None:
            raise credentials_exception
        token_data = TokenData(email=email, role=role)
    except JWTError:
        raise credentials_exception
    user = get_user(email=token_data.email)
    if user is None:
        raise credentials_exception
    return user

async def get_current_admin(token: str = Depends(oauth2_scheme)) -> dict:
    """Ensure the current user is an Admin."""
    user = await get_current_user(token)
    if user["role_name"] != "Admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized: Admin access required"
        )
    return user

# Authentication Endpoints
@app.post("/signup", tags=["Authentication"])
async def signup(user: UserCreate):
    """Register a new user with role_id (1 for Admin, 2 for User, defaults to 2)."""
    role = get_role(user.role_id)
    if not role:
        raise HTTPException(status_code=400, detail="Invalid role_id. Must be 1 (Admin) or 2 (User).")
    if get_user(user.email):
        raise HTTPException(status_code=400, detail="Email already registered")
    hashed_password = hash_password(user.password)
    user_dict = {
        "first_name": user.first_name,
        "last_name": user.last_name,
        "email": user.email,
        "hashed_password": hashed_password,
        "role_id": user.role_id,
        "role_name": role["role_name"]
    }
    try:
        users_collection.insert_one(user_dict)
    except DuplicateKeyError:
        raise HTTPException(status_code=400, detail="Email already registered")
    return {"message": "User registered successfully"}

@app.post("/login", response_model=Token, tags=["Authentication"])
async def login(form_data: OAuth2PasswordRequestForm = Depends()):
    """Authenticate user and return JWT token."""
    user = get_user(form_data.username)  # form_data.username is the email
    if not user or not verify_password(form_data.password, user["hashed_password"]):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect email or password",
            headers={"WWW-Authenticate": "Bearer"},
        )

    access_token = create_access_token(
        data={"sub": user["email"], "role": user["role_name"]}
    )

    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": {
            "_id": str(user["_id"]),
            "first_name": user["first_name"],
            "last_name": user["last_name"],
            "email": user["email"],
            "role_id": user["role_id"],
            "role_name": user["role_name"],
        }
    }


from bson import ObjectId
from fastapi import APIRouter, Depends

@app.get("/users", tags=["Authentication"])
async def get_all_users(current_admin: dict = Depends(get_current_admin)):
    """
    Get all registered users.
    Only accessible by Admin.
    """
    users = list(users_collection.find({}, {"hashed_password": 0}))  

    # Convert ObjectId to string
    for user in users:
        user["_id"] = str(user["_id"])

    return {"users": users}

from fastapi import FastAPI, Depends, HTTPException
from bson import ObjectId

@app.delete("/users/{user_id}", tags=["Authentication"])
async def delete_user(user_id: str, current_admin: dict = Depends(get_current_admin)):
    """
    Delete a user by ID.
    Only accessible by Admin.
    """
    # Validate ObjectId
    if not ObjectId.is_valid(user_id):
        raise HTTPException(status_code=400, detail="Invalid user ID format")

    result = users_collection.delete_one({"_id": ObjectId(user_id)})

    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")

    return {"message": "User deleted successfully", "user_id": user_id}



from bson import ObjectId

class UpdateUserRole(BaseModel):
    user_id: str
    role_id: int

@app.put("/users/update-role", tags=["Authentication"])
async def update_user_role(
    data: UpdateUserRole,
    current_admin: dict = Depends(get_current_admin)
    ):
    """
    Update a user's role (Admin-only).
    Requires user_id and new role_id.
    """
    # Validate role
    role = get_role(data.role_id)
    if not role:
        raise HTTPException(status_code=400, detail="Invalid role_id. Must be 1 (Admin) or 2 (User).")

    # Convert string user_id to ObjectId
    try:
        user_obj_id = ObjectId(data.user_id)
    except:
        raise HTTPException(status_code=400, detail="Invalid user_id format")

    # Find user
    user = users_collection.find_one({"_id": user_obj_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Update role
    users_collection.update_one(
        {"_id": user_obj_id},
        {"$set": {"role_id": data.role_id, "role_name": role["role_name"]}}
    )

    return {"message": f"User role updated to {role['role_name']} successfully"}


#--------------------User-Analysis------------------------
@app.get("/users/{user_id}/analysis", tags=["User Analysis"])
async def get_user_analysis(user_id: str):
    """
    Fetch all past analysis records for a given user_id.
    """
    try:
        # Validate ObjectId format (in case user_id is Mongo _id)
        # but since you are storing user_id as string, we can just use it directly
        records = list(user_analysis.find({"user_id": user_id}, {"_id": 0}))
        
        if not records:
            return {"message": "No analysis records found for this user", "data": []}

        return {"user_id": user_id, "data": records}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching analysis: {str(e)}")


#----------------------------get the contracts_cache--------------------------------------------
from fastapi import HTTPException
from bson import ObjectId

# ------------------ Contract Cache ------------------
@app.get("/contracts-cache/{cache_id}", tags=["Contracts Cache"])
async def get_contract_cache(cache_id: str):
    """
    Fetch a cached contract analysis by its ID.
    """
    try:
        record = contracts_cache.find_one({"_id": cache_id}, {"_id": 0})
        if not record:
            raise HTTPException(status_code=404, detail="Contract cache not found")

        return {"cache_id": cache_id, "data": record}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching contract cache: {str(e)}")

#--------------------------------download-contract-------------------------------------
@app.get("/download-contract/{cache_id}", tags=["Contracts-Cache"])
async def download_contract(cache_id: str):
    """
    Download the first uploaded contract PDF from uploaded_files/{cache_id}.
    """
    try:
        contract_dir = UPLOAD_DIR / cache_id
        print(f"Looking in: {contract_dir}")  # DEBUG

        if not contract_dir.exists() or not contract_dir.is_dir():
            raise HTTPException(status_code=404, detail="No folder found for this cache_id")

        # Get list of files in the folder
        files = list(contract_dir.glob("*"))
        print(f"Files found: {files}")  # DEBUG

        if not files:
            raise HTTPException(status_code=404, detail="No files found in this cache_id folder")

        # Pick the first file (could sort if you want deterministic order)
        pdf_path = files[0]
        print(f"Returning file: {pdf_path}")  # DEBUG
        print("--------------------------------------------------->" + os.path.basename(str(pdf_path)))

        return FileResponse(
            pdf_path,
            media_type="application/pdf",
            filename=os.path.basename(str(pdf_path))   # ✅ sends original filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()  # show full error in logs
        raise HTTPException(status_code=500, detail=f"Error fetching PDF: {str(e)}")


#----------------------------Excel Report Download-------------------------------------------------------------------
from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import StreamingResponse
from pymongo import MongoClient
from bson import ObjectId
import pandas as pd
import io
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Color

# Assuming `db` is already defined and connected to MongoDB
contracts_collection = db["contracts_cache"]
clauses_collection = db["clauses"]

# ------------------ Excel Formatting ------------------
def format_sheet(writer, sheet_name: str):
    worksheet = writer.sheets[sheet_name]
    
    # Define color styles
    metadata_fill = PatternFill(start_color="0F6368", end_color="0F6368", fill_type="solid")
    table_fill = PatternFill(start_color="299E66", end_color="299E66", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)  # White and bold for headers
    white_font_regular = Font(color="FFFFFF")  # White for regular text
    
    # Apply formatting to metadata section (rows 1-3)
    for row in worksheet[1:3]:  # Rows are 1-based in openpyxl
        for cell in row:
            cell.fill = metadata_fill
            cell.font = white_font_regular
    
    # Apply formatting to table header (row 5, since startrow=4 in df_table.to_excel)
    for cell in worksheet[5]:  # Row 5 is the table header
        cell.fill = table_fill
        cell.font = white_font  # Bold white font for headers
    
    # Apply formatting to table data (rows 6 and below)
    for row in worksheet.iter_rows(min_row=6, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.fill = table_fill
            cell.font = white_font_regular
    
    # Adjust column widths based on content
    for col_idx, col in enumerate(worksheet.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = max_length + 2

# ------------------ Download Excel Endpoint ------------------
@app.get("/download-excel", tags=["Contracts-Cache"])
def download_excel_format(_id: str = Query(...)):
    # Fetch the contract analysis document
    document = contracts_collection.find_one({"_id": _id}) or contracts_collection.find_one({"_id": ObjectId(_id)})
    if not document:
        raise HTTPException(status_code=404, detail="Analysis not found")

    # Contract metadata
    contract_name = document.get("pdf_name", "Contract.pdf")
    num_clauses = len(document.get("clauses", []))
    description = document.get("analysis", {}).get("description", "")

    # Loop over clauses in the contract
    table_rows = []
    for clause_id in document.get("clauses", []):
        # Fetch clause document
        clause_doc = clauses_collection.find_one({"_id": clause_id}) or clauses_collection.find_one({"_id": ObjectId(clause_id)})
        clause_name = clause_doc.get("clause_name") if clause_doc else ""
        clause_text = clause_doc.get("clause_text") if clause_doc else ""

        # Match analysis for this clause (safe dict check)
        matched_analysis = next(
            (
                a for a in document.get("analysis", {}).get("analysis", [])
                if isinstance(a, dict) and a.get("compliance_area", "").lower() == clause_name.lower()
            ),
            {}
        )

        table_rows.append({
            "Clause": clause_name,
            "Clause Text": clause_text,
            "Status": matched_analysis.get("missing_clause", ""),
            "Reason": matched_analysis.get("reason", ""),
            "Extracted Clause": matched_analysis.get("extracted_text", "")
        })

    # Create DataFrame
    df_table = pd.DataFrame(table_rows, columns=["Clause", "Clause Text", "Status", "Reason", "Extracted Clause"])

    # Write Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Add top metadata manually
        meta_df = pd.DataFrame({
            0: ["Contract Name", "No of Clauses", "Description"],
            1: [contract_name, num_clauses, description]
        })
        meta_df.to_excel(writer, index=False, header=False, startrow=0)
        # Add table starting from row 4
        df_table.to_excel(writer, index=False, startrow=4)
        format_sheet(writer, "Sheet1")

    output.seek(0)
    filename = f"{contract_name}_report.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )